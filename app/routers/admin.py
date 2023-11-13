from datetime import datetime
from uuid import uuid4

import urllib.parse
from fastapi import APIRouter, Depends, Response, Query, HTTPException

import asyncio
import openpyxl
import io
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from app.utils.database import get_db, check_complete_task, check_complete_post
from app.utils.jwt_handler import get_current_user
from app.models.models import Project, Task
from app.utils.database import compare
import httpx

router = APIRouter(
    prefix="/admin",
    tags=["Admin и возможности"]
)

@router.get("/get_users_list/")
async def get_admin_panel(current_user: str = Depends(get_current_user)):
    db = get_db()
    users_collection = db['users']
    user_list = await users_collection.find({'role': 0}, {'_id': 0, 'username': 1}).to_list(length=None)
    return user_list

@router.get("/show_user_tasks&posts/{username}")
async def show_collection(username: str, current_user: str = Depends(get_current_user)):
    db = get_db()
    task_collection = db['tasks']
    posts_collection = db['posts']
    project_collection = db['projects']
    user_tasks = await task_collection.find({"members": username, 'status': 'current'}, {'_id': 0,'id': 1,'name': 1, 'description': 1, 'project': 1,'deadline':1,'comments':1,'created_at': 1, 'status': 1, 'time_completed': 1, 'type': 1,'members': 1,'status':1}).to_list(length=None)
    user_posts = await posts_collection.find({"member": username, 'status': 'current'}, {'_id': 0,'id': 1,'name': 1, 'description': 1, 'member': 1,'created_at': 1, 'status': 1, 'time_completed': 1, 'deadline': 1, 'task': 1,'type':1,'action':1,'client':1,'status':1}).to_list(length=None)

    for user_task in user_tasks:
        project_id = user_task['project']
        project_id = urllib.parse.unquote(project_id)
        project_name = await project_collection.find_one({'id': project_id}, {'name': 1})
        user_task["project_name"] = project_name['name'] if project_name else "Неизвестный проект"
        user_task["project"] = str(project_id)  # Преобразование ObjectId в строку
    
    for user_post in user_posts:
        task_id = user_post['task']
        task_id = urllib.parse.unquote(task_id)
        task_name = await task_collection.find_one({'id': task_id}, {'name': 1})
        user_post["task_name"] = task_name['name'] if task_name else "Неизвестная задача"
        user_post["task"] = str(task_id)  # Преобразование ObjectId в строку
    
    completed_lst = user_posts + user_tasks
    return completed_lst

@router.post("/create_project")
async def create_project(project: Project, current_user: str = Depends(get_current_user)):
    db = get_db()
    project_collection = db['projects']
    project_dict = project.dict()
    project_dict['id'] = str(uuid4())
    project_dict['status'] = 'current'
    timestamp = datetime.now().timestamp()
    timestamp_without_ms = round(timestamp)
    project_dict['created_at'] = timestamp_without_ms
    project_name = project.name
    new_project = await project_collection.insert_one(project_dict)
    data_to_sed = {
        "project_name": project_name
    }

    if new_project:
        return {"message": "New project is created"} 
    async with httpx.AsyncClient() as client:
        response = await client.post("https://test.4dev.kz/gsd_dev/hs/ServiceDesk1C/Tasks", json=data_to_sed)
        if response.status_code == 200:
            response_data = response.json()
            print(response_data)

            await project_collection.update_one(
                {"id": project_dict['id']},
                {"$set": {"1C_id": response_data}}
            )

            return {"message": "New project is created and response data is stored."}
        else:
            return {"error": "Failed to send data to the other endpoint."}
    
@router.post("/create_task")
async def create_task(task: Task, current_user: str = Depends(get_current_user)):
    db = get_db()
    task_collection = db['tasks']
    project_collection = db['projects']
    project_id = await project_collection.find_one({'name': task.project},{'_id': 0, 'id': 1})
    task.project = project_id['id']
    task_dict = task.dict()
    task_dict['id'] = str(uuid4())
    task_dict['type'] = "task"
    task_dict['status'] = 'current'
    timestamp = datetime.now().timestamp()
    timestamp_without_ms = round(timestamp)
    task_dict['created_at'] = timestamp_without_ms
    compared_lst = await compare(task_dict['project'], "projects")

    project_found = False

    for id in compared_lst:
        if id['id'] == task_dict['project']:
            project_found = True
            break

    if project_found:
        new_task = await task_collection.insert_one(task_dict)
        if new_task:
            return {"message": "New task is created"}
    else:
        raise HTTPException(status_code=404, detail="Project not found")

@router.post("/exportexcel/all_posts")
async def export_all_posts(start: float = Query(default=0, ge=0), end: float = Query(default=2000000000, le=2000000000), current_user: str = Depends(get_current_user)):
    try:
        db = get_db()
        posts_collection = db['posts']
        task_collection = db['tasks']
        project_collection = db['projects']
        posts = await posts_collection.find({"created_at":{"$gte": start, "$lte": end}}, {'_id': 0, 'member': 1, 'name': 1, 'description': 1, 'created_at': 1,'task':1,'status':1}).to_list(length=None)
        tasks = []
        for data in posts:
            task_id = data.get('task')
            if task_id:
                tasks.append(task_collection.find_one({'id': task_id}, {'_id': 0, 'project': 1}))
        
        project_ids = await asyncio.gather(*tasks)

        for data, project_id in zip(posts, project_ids):
            if project_id:
                project_info = await project_collection.find_one({'id': project_id['project']}, {'_id': 0, 'name': 1})
                data['project_name'] = project_info.get('name', '')
            else:
                data['project_name'] = ''

        posts = sorted(posts, key=lambda x: x['member'])
        book = openpyxl.Workbook()
        sheet = book.active

        sheet['A1'] = "СОТРУДНИК"
        sheet['B1'] = "ЗАГОЛОВОК"
        sheet['C1'] = "ИНФОРМАЦИЯ"
        sheet['D1'] = "СТАТУС"
        sheet['E1'] = "ПРОЕКТ"
        sheet['F1'] = "ТРУДОЗАТРАТЫ"
        sheet['G1'] = "ДАТА СОЗДАНИЯ"

        for cell in sheet[1]:
                cell.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")

        row = 2
        for data in posts:
            timestamp = data['created_at']
            data_post = datetime.fromtimestamp(timestamp)
            data_post_without_ms = data_post.replace(microsecond=0)
            sheet[row][0].value = data['member']
            sheet[row][1].value = data['name']
            sheet[row][2].value = data['description']
            sheet[row][3].value = 'В работе' if data['status'] == 'current' else 'Завершено'
            sheet[row][4].value = data['project_name']
            sheet[row][5].value = '0 часов'
            sheet[row][6].value = data_post_without_ms
            row += 1

        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length) * 1.2
            if adjusted_width < 18:
                adjusted_width = 18
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        output = io.BytesIO()

        book.save(output)
        output.seek(0)

        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=all_posts.xlsx"})
    except Exception as error: 
        return "something wrong"
    
@router.post("/exportexcel/{username}")
async def export_user_posts(username: str, start: float = Query(default=0, ge=0), end: float = Query(default=2000000000, le=2000000000), current_user: str = Depends(get_current_user)):
    try:
        db = get_db()
        posts_collection = db['posts']
        task_collection = db['tasks']
        project_collection = db['projects']
        posts = await posts_collection.find({"member": username, "created_at": {"$gte": start, "$lte": end}}, {'_id': 0, 'member': 1, 'name': 1, 'description': 1, 'created_at': 1,'task':1,'status':1}).to_list(length=None)
        tasks = []
        for data in posts:
            task_id = data.get('task')
            if task_id:
                tasks.append(task_collection.find_one({'id': task_id}, {'_id': 0, 'project': 1}))
        
        project_ids = await asyncio.gather(*tasks)

        for data, project_id in zip(posts, project_ids):
            if project_id:
                project_info = await project_collection.find_one({'id': project_id['project']}, {'_id': 0, 'name': 1})
                data['project_name'] = project_info.get('name', '')
            else:
                data['project_name'] = ''

        posts = sorted(posts, key=lambda x: x['member'])
        book = openpyxl.Workbook()
        sheet = book.active

        sheet['A1'] = "СОТРУДНИК"
        sheet['B1'] = "ЗАГОЛОВОК"
        sheet['C1'] = "ИНФОРМАЦИЯ"
        sheet['D1'] = "СТАТУС"
        sheet['E1'] = "ПРОЕКТ"
        sheet['F1'] = "ТРУДОЗАТРАТЫ"
        sheet['G1'] = "ДАТА СОЗДАНИЯ"

        for cell in sheet[1]:
                cell.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")

        row = 2
        for data in posts:
            timestamp = data['created_at']
            data_post = datetime.fromtimestamp(timestamp)
            data_post_without_ms = data_post.replace(microsecond=0)
            sheet[row][0].value = data['member']
            sheet[row][1].value = data['name']
            sheet[row][2].value = data['description']
            sheet[row][3].value = 'В работе' if data['status'] == 'current' else 'Завершено'
            sheet[row][4].value = data['project_name']
            sheet[row][5].value = '0 часов'
            sheet[row][6].value = data_post_without_ms
            row += 1

        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length) * 1.2
            if adjusted_width < 18:
                adjusted_width = 18
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        output = io.BytesIO()

        book.save(output)
        output.seek(0)

        return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=all_posts.xlsx"})
    except Exception as error: 
        return "something wrong"

@router.put("/change_project/{project_id}")
async def change_project(project_id: str, project: Project, current_user: str = Depends(get_current_user)):
    db = get_db()
    project_collection = db['projects']
    old_project = await project_collection.find_one({'id': project_id})
    if old_project:
        if old_project['deadline'] != project.deadline:
            if 'archive_deadline' not in old_project:
                old_project['archive_deadline'] = []
            old_project['archive_deadline'].append(old_project['deadline'])
            old_project['deadline'] = project.deadline
        old_project['name'] = project.name
        old_project['description'] = project.description
        old_project['members'] = project.members
        await project_collection.replace_one({'id': project_id}, old_project)
    else:
        return {'message': 'Project not found'}
    
@router.put("/change_task/{task_id}")
async def change_task(task_id: str, task: Task, current_user: str = Depends(get_current_user)):
    db = get_db()
    task_collection = db['tasks']
    project_collection = db['projects']
    old_task = await task_collection.find_one({'id': task_id})
    old_project_name = await project_collection.find_one({'id': old_task['project']}, {'_id': 0, 'name': 1})
    new_project_id = await project_collection.find_one({'name': task.project}, {'_id':0, 'id': 1})
    if old_task:
        if old_task['deadline'] != task.deadline:
            if 'archive_deadline' not in old_task:
                old_task['archive_deadline'] = []
            old_task['archive_deadline'].append(old_task['deadline'])
            old_task['deadline'] = task.deadline
        old_task['name'] = task.name
        old_task['description'] = task.description
        old_task['members'] = task.members
        if old_project_name['name'] != task.project:
            old_task['project'] = new_project_id['id']
        await task_collection.replace_one({'id': task_id}, old_task)
        return {'message': 'Task updated'}
    else:
        return {'message': 'Task not found'}


@router.patch("/complete_project/{project_id}")
async def complete_project(project_id: str, current_user: str = Depends(get_current_user)):
    db = get_db()
    project_collection = db['projects']
    project = await project_collection.find_one({'id': project_id})
    checked_tasks = await check_complete_task(project['id'])
    print(checked_tasks)
    
    da_ili_net = True
    if len(checked_tasks) == 0:
        da_ili_net = False
    for name in checked_tasks:
        if name['status'] == 'current':
            da_ili_net = False
            break

    if da_ili_net:
        result = await project_collection.update_one({'id': project_id}, {'$set': {'status': 'completed'}})
        if result.matched_count == 1:
            timestamp = datetime.now().timestamp()
            timestamp_without_ms = round(timestamp)
            await project_collection.update_one({'id': project_id},{'$set':{'time_completed': timestamp_without_ms}})
        else:
            raise HTTPException(status_code=404, detail="Item not found")
    else:
        raise HTTPException(status_code=404, detail="You have 'status': 'current'")
    

@router.patch("/complete_task/{task_id}")
async def complete_task(task_id: str, current_user: str = Depends(get_current_user)):
    db = get_db()
    task_collection = db['tasks']
    task = await task_collection.find_one({'id': task_id})
    checked_posts = await check_complete_post(task['id'])

    da_ili_net = True
    if len(checked_posts) == 0:
        da_ili_net = False
    for name in checked_posts:
        if name['status'] == 'current':
            da_ili_net = False
            break
    if da_ili_net:
        result = await task_collection.update_one({'id': task_id}, {'$set': {'status': 'completed'}})
        if result.matched_count == 1:
            timestamp = datetime.now().timestamp()
            timestamp_without_ms = round(timestamp)
            await task_collection.update_one({'id': task_id},{'$set':{'time_completed': timestamp_without_ms}})
        else:
            raise HTTPException(status_code=404, detail="Item not found")
    else:
        raise HTTPException(status_code=404, detail="You have 'status': 'current'")



@router.delete("/delete_project/{project_id}")
async def delete_project(project_id: str, current_user: str = Depends(get_current_user)):
    db = get_db()
    project_collection = db['projects']
    task_collection = db['tasks']
    post_collection = db['posts']
    project = await project_collection.find_one({'id': project_id}, {'_id': 0, 'id': 1})
    project_id = project['id']
    tasks_lst = await task_collection.find({'project': project_id}, {'_id': 0, 'id': 1}).to_list(length=None)
    
    for task in tasks_lst:
        id = task['id']
        await post_collection.delete_many({'task': id})

    result_task = await task_collection.delete_many({'project': project_id})
    if result_task:
        result = await project_collection.delete_one({'id': project_id})
        if result:
            return {"message": "Project and his tasks & posts have been deleted"}   
        

@router.delete("/delete_task/{task_id}")
async def delete_task(task_id: str, current_user: str = Depends(get_current_user)):
    db = get_db()
    task_collection = db['tasks']
    post_collection = db['posts']
    posts_lst = await post_collection.find({'task': task_id}, {'_id': 0, 'id': 1}).to_list(length=None)

    for post in posts_lst:
        id = post['id']
        await post_collection.delete_one({'task': id})

    result = await task_collection.delete_one({'id': task_id})
    if result:
        return {"message": "Task and his posts have been deleted"} 